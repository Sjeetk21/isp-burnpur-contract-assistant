from __future__ import annotations

import argparse
import json
import math
import os
import re
import time
import urllib.error
import urllib.request
from dataclasses import asdict, dataclass
from http.server import BaseHTTPRequestHandler, ThreadingHTTPServer
from pathlib import Path
from typing import Any


@dataclass
class Config:
    source_dir: str = os.environ.get("SOURCE_DOC_DIR", r"C:\Users\INP\Downloads")
    index_dir: str = "index"
    child_word_limit: int = 220
    child_overlap: int = 40
    model: str = os.environ.get("GEMINI_MODEL", "gemini-2.5-flash")
    top_k: int = 8


SUPPORTED = {".docx", ".xlsx", ".xls", ".pdf", ".txt", ".md", ".csv"}
QUERY_EXPANSION = {
    "wr": "work roll",
    "bur": "backup roll",
    "hsm": "hot strip mill",
    "fce": "furnace",
    "mtpa": "million tonnes per annum",
    "tph": "tons per hour",
    "qap": "quality assurance plan",
    "ic": "inspection certificate",
}
ROUTE_TERMS = {
    "capacity": ["capacity", "mtpa", "tph", "million tonnes"],
    "safety": ["safety", "hazard", "ppe", "fire", "emergency"],
    "inspection": ["inspection", "qap", "witness", "review", "test certificate"],
    "scope": ["scope", "supply", "battery limit", "exclusion", "interface"],
    "dimension": ["width", "thickness", "diameter", "length", "mm"],
    "payment": ["payment", "billing", "invoice", "price"],
}


def norm(text: Any) -> str:
    return re.sub(r"\s+", " ", "" if text is None else str(text).replace("\u00a0", " ")).strip()


def tokens(text: str) -> list[str]:
    return re.findall(r"[a-z0-9.]+", text.lower())


def chunk_words(words: list[str], size: int, overlap: int) -> list[str]:
    if not words:
        return []
    chunks = []
    step = max(size - overlap, 1)
    for start in range(0, len(words), step):
        part = words[start : start + size]
        if part:
            chunks.append(" ".join(part))
        if start + size >= len(words):
            break
    return chunks


def parse_docx(path: Path) -> list[dict[str, Any]]:
    from docx import Document

    out = []
    section = "General"
    doc = Document(path)
    for para in doc.paragraphs:
        text = norm(para.text)
        if not text:
            continue
        first = text.split()[0] if text.split() else ""
        style = para.style.name if para.style else ""
        if (style.startswith("Heading") or re.match(r"^\d+(\.\d+)*\.?$", first)) and len(text) < 180:
            section = text
            continue
        out.append(make_record(path, text, section, None, "docx"))
    return out


def parse_xlsx(path: Path) -> list[dict[str, Any]]:
    from openpyxl import load_workbook

    out = []
    wb = load_workbook(path, read_only=True, data_only=True)
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        for row_no, row in enumerate(ws.iter_rows(values_only=True), 1):
            cells = [norm(cell) for cell in row if norm(cell)]
            if len(cells) >= 2:
                out.append(make_record(path, " | ".join(cells), f"Sheet: {sheet}", row_no, "xlsx"))
    wb.close()
    return out


def parse_pdf(path: Path) -> list[dict[str, Any]]:
    try:
        from pypdf import PdfReader
    except Exception:
        return []
    out = []
    reader = PdfReader(str(path))
    for page_no, page in enumerate(reader.pages, 1):
        text = norm(page.extract_text() or "")
        if text:
            out.append(make_record(path, text, f"PDF page {page_no}", page_no, "pdf"))
    return out


def parse_text(path: Path) -> list[dict[str, Any]]:
    return [make_record(path, norm(path.read_text(encoding="utf-8", errors="ignore")), "Text", None, path.suffix.lower().strip("."))]


def make_record(path: Path, text: str, section: str, page: int | None, source_type: str) -> dict[str, Any]:
    return {
        "text": text,
        "source_file": path.name,
        "chapter": path.stem.split("_")[0].replace("-", " "),
        "section": section,
        "page_number": page,
        "source_type": source_type,
    }


def source_files(root: str) -> list[Path]:
    base = Path(root)
    return sorted(
        p for p in base.rglob("*")
        if p.is_file() and not p.name.startswith("~$") and p.suffix.lower() in SUPPORTED
    )


def load_records(config: Config) -> list[dict[str, Any]]:
    parsers = {
        ".docx": parse_docx,
        ".xlsx": parse_xlsx,
        ".xls": parse_xlsx,
        ".pdf": parse_pdf,
        ".txt": parse_text,
        ".md": parse_text,
        ".csv": parse_text,
    }
    records = []
    for file_path in source_files(config.source_dir):
        try:
            parsed = parsers[file_path.suffix.lower()](file_path)
            records.extend(parsed)
            print(f"[OK] {file_path.name}: {len(parsed)} records")
        except Exception as exc:
            print(f"[WARN] {file_path}: {exc}")
    return records


def build_index(config: Config) -> dict[str, Any]:
    index_path = Path(config.index_dir)
    index_path.mkdir(parents=True, exist_ok=True)
    records = load_records(config)
    children = []
    route_index = {key: [] for key in ROUTE_TERMS}
    doc_freq: dict[str, int] = {}
    lengths = []

    for parent_id, rec in enumerate(records):
        chunks = chunk_words(rec["text"].split(), config.child_word_limit, config.child_overlap)
        for chunk_no, text in enumerate(chunks):
            child = {
                "id": len(children),
                "parent_id": parent_id,
                "chunk_index": chunk_no,
                "text": text,
                "metadata": {k: rec[k] for k in rec if k != "text"},
            }
            children.append(child)
            tok_set = set(tokens(text))
            lengths.append(len(tokens(text)))
            for tok in tok_set:
                doc_freq[tok] = doc_freq.get(tok, 0) + 1
            lower = text.lower()
            for route, terms in ROUTE_TERMS.items():
                if any(term in lower for term in terms):
                    route_index[route].append(child["id"])

    payload = {
        "config": asdict(config),
        "records": records,
        "children": children,
        "doc_freq": doc_freq,
        "avg_len": sum(lengths) / max(len(lengths), 1),
        "route_index": route_index,
        "built_at": time.strftime("%Y-%m-%d %H:%M:%S"),
    }
    Path(config.index_dir, "contract_index.json").write_text(json.dumps(payload, ensure_ascii=False), encoding="utf-8")
    summary = {
        "source_dir": Path(config.source_dir).name,
        "source_files": len(source_files(config.source_dir)),
        "records": len(records),
        "children": len(children),
        "index_file": "index/contract_index.json",
    }
    Path(config.index_dir, "summary.json").write_text(json.dumps(summary, indent=2), encoding="utf-8")
    return summary


class ContractIndex:
    def __init__(self, index_dir: str = "index"):
        payload = json.loads(Path(index_dir, "contract_index.json").read_text(encoding="utf-8"))
        self.children = payload["children"]
        self.doc_freq = payload["doc_freq"]
        self.avg_len = payload["avg_len"]
        self.route_index = payload["route_index"]
        self.n_docs = max(len(self.children), 1)

    def expand_query(self, query: str) -> str:
        expanded = query
        lower = query.lower()
        for short, full in QUERY_EXPANSION.items():
            if re.search(rf"\b{re.escape(short)}\b", lower):
                expanded += " " + full
        return expanded

    def route_bonus_ids(self, query: str) -> set[int]:
        lower = query.lower()
        ids = set()
        for route, terms in ROUTE_TERMS.items():
            if route in lower or any(term in lower for term in terms):
                ids.update(self.route_index.get(route, [])[:80])
        return ids

    def search(self, query: str, top_k: int = 8) -> list[dict[str, Any]]:
        query = self.expand_query(query)
        q_tokens = tokens(query)
        bonus_ids = self.route_bonus_ids(query)
        k1, b = 1.5, 0.75
        scored = []
        for child in self.children:
            text_tokens = tokens(child["text"])
            if not text_tokens:
                continue
            freq = {}
            for tok in text_tokens:
                freq[tok] = freq.get(tok, 0) + 1
            score = 0.0
            for tok in q_tokens:
                if tok not in freq:
                    continue
                df = self.doc_freq.get(tok, 0)
                idf = math.log(1 + (self.n_docs - df + 0.5) / (df + 0.5))
                denom = freq[tok] + k1 * (1 - b + b * len(text_tokens) / max(self.avg_len, 1))
                score += idf * (freq[tok] * (k1 + 1)) / denom
            if child["id"] in bonus_ids:
                score += 1.5
            if score > 0:
                scored.append({**child, "score": score})
        return sorted(scored, key=lambda item: item["score"], reverse=True)[:top_k]


def citation_line(item: dict[str, Any], idx: int) -> str:
    meta = item["metadata"]
    page = f", page/row {meta['page_number']}" if meta.get("page_number") else ""
    return f"{idx}. {meta['source_file']} | {meta['section']}{page}"


def gemini_answer(question: str, hits: list[dict[str, Any]], model: str) -> str:
    api_key = os.environ.get("GEMINI_API_KEY") or os.environ.get("GOOGLE_API_KEY")
    context = "\n\n".join(
        f"--- SEGMENT {idx} ---\nSOURCE: {citation_line(hit, idx)}\n{hit['text']}"
        for idx, hit in enumerate(hits, 1)
    )
    citations = "\n".join(citation_line(hit, idx) for idx, hit in enumerate(hits, 1))
    if not api_key:
        return "Set GEMINI_API_KEY or GOOGLE_API_KEY to generate an answer.\n\nRetrieved evidence:\n" + citations

    prompt = (
        "You are a Principal Contract AI Assistant for the ISP Burnpur Hot Strip Mill project. "
        "Answer only from the verified contract segments. Preserve exact numbers, clause references, "
        "technical parameters, responsibilities, and exceptions. If the context is insufficient, say so.\n\n"
        f"CONTRACT SEGMENTS:\n{context}\n\nQUESTION: {question}\n\nANSWER:"
    )
    body = json.dumps({
        "contents": [{"parts": [{"text": prompt}]}],
        "generationConfig": {"temperature": 0.0, "maxOutputTokens": 1400},
    }).encode("utf-8")
    url = f"https://generativelanguage.googleapis.com/v1beta/models/{model}:generateContent?key={api_key}"
    req = urllib.request.Request(url, data=body, headers={"Content-Type": "application/json"})
    try:
        with urllib.request.urlopen(req, timeout=60) as resp:
            payload = json.loads(resp.read().decode("utf-8"))
        text = payload["candidates"][0]["content"]["parts"][0]["text"]
        return f"{text}\n\nOfficial contract references:\n{citations}"
    except urllib.error.HTTPError as exc:
        return f"Gemini API error {exc.code}: {exc.read().decode('utf-8', errors='ignore')}\n\nRetrieved evidence:\n{citations}"
    except Exception as exc:
        return f"Gemini request failed: {exc}\n\nRetrieved evidence:\n{citations}"


def ask(question: str, config: Config) -> dict[str, Any]:
    index = ContractIndex(config.index_dir)
    hits = index.search(question, config.top_k)
    answer = gemini_answer(question, hits, config.model) if hits else "No matching contract evidence was found."
    return {"answer": answer, "hits": hits}


HTML = """<!doctype html>
<html>
<head>
  <meta charset="utf-8">
  <title>ISP Burnpur Contract Assistant</title>
  <style>
    body { font-family: Segoe UI, Arial, sans-serif; margin: 0; background: #f6f7f9; color: #1f2937; }
    header { background: #123; color: white; padding: 18px 28px; }
    main { max-width: 1100px; margin: 0 auto; padding: 24px; }
    textarea { width: 100%; min-height: 80px; padding: 12px; font-size: 15px; box-sizing: border-box; }
    button { margin-top: 10px; padding: 10px 16px; background: #0f5c8c; color: white; border: 0; border-radius: 4px; cursor: pointer; }
    pre { white-space: pre-wrap; background: white; padding: 16px; border: 1px solid #d7dce2; border-radius: 6px; }
    .hit { background: white; border: 1px solid #d7dce2; border-radius: 6px; padding: 12px; margin: 12px 0; }
    .muted { color: #596579; font-size: 13px; }
  </style>
</head>
<body>
<header><h2>ISP Burnpur Contract Assistant</h2><div>Local evidence search with Gemini-grounded answers</div></header>
<main>
  <textarea id="q" placeholder="Ask about safety, scope, inspection, capacity, exclusions, payment..."></textarea>
  <br><button onclick="ask()">Ask Contract</button>
  <h3>Answer</h3><pre id="answer">Ready.</pre>
  <h3>Evidence</h3><div id="hits"></div>
</main>
<script>
async function ask() {
  const question = document.getElementById('q').value;
  document.getElementById('answer').textContent = 'Searching...';
  const res = await fetch('/ask', {method:'POST', headers:{'Content-Type':'application/json'}, body: JSON.stringify({question})});
  const data = await res.json();
  document.getElementById('answer').textContent = data.answer;
  document.getElementById('hits').innerHTML = data.hits.map((h, i) => `<div class="hit"><b>${i+1}. ${h.metadata.source_file}</b><div class="muted">${h.metadata.section} | page/row ${h.metadata.page_number || ''} | score ${h.score.toFixed(3)}</div><div>${h.text.substring(0,1200)}</div></div>`).join('');
}
</script>
</body>
</html>"""


class Handler(BaseHTTPRequestHandler):
    config = Config()

    def _send(self, status: int, body: bytes, content_type: str) -> None:
        self.send_response(status)
        self.send_header("Content-Type", content_type)
        self.send_header("Content-Length", str(len(body)))
        self.end_headers()
        self.wfile.write(body)

    def do_GET(self) -> None:
        if self.path == "/":
            self._send(200, HTML.encode("utf-8"), "text/html; charset=utf-8")
        elif self.path == "/summary":
            body = Path(self.config.index_dir, "summary.json").read_bytes()
            self._send(200, body, "application/json")
        else:
            self._send(404, b"Not found", "text/plain")

    def do_POST(self) -> None:
        if self.path != "/ask":
            self._send(404, b"Not found", "text/plain")
            return
        length = int(self.headers.get("Content-Length", "0"))
        payload = json.loads(self.rfile.read(length).decode("utf-8"))
        result = ask(payload.get("question", ""), self.config)
        self._send(200, json.dumps(result, ensure_ascii=False).encode("utf-8"), "application/json; charset=utf-8")


def serve(config: Config, host: str, port: int) -> None:
    Handler.config = config
    server = ThreadingHTTPServer((host, port), Handler)
    print(f"Open http://{host}:{port}")
    server.serve_forever()


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("command", choices=["build", "ask", "serve"])
    parser.add_argument("--source-dir", default=Config().source_dir)
    parser.add_argument("--index-dir", default=Config().index_dir)
    parser.add_argument("--question", default="")
    parser.add_argument("--host", default="127.0.0.1")
    parser.add_argument("--port", type=int, default=7860)
    parser.add_argument("--top-k", type=int, default=8)
    args = parser.parse_args()
    config = Config(source_dir=args.source_dir, index_dir=args.index_dir, top_k=args.top_k)
    if args.command == "build":
        print(json.dumps(build_index(config), indent=2))
    elif args.command == "ask":
        print(json.dumps(ask(args.question, config), indent=2, ensure_ascii=False))
    else:
        serve(config, args.host, args.port)


if __name__ == "__main__":
    main()
